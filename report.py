import csv
import json
import os
import re
from datetime import datetime
from typing import List, Dict, Set

import common
from cve import CVEEnricher
from html_report import export_html

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None

JSON_SCHEMA = {
    "$id": "https://nmapx.local/schema/v1",
    "$schema": "http://json-schema.org/draft-07/schema#",
    "title": "NmapX Scan Report",
    "type": "object",
    "properties": {
        "generated_at": {"type": "string", "format": "date-time"},
        "tool": {"type": "string"},
        "schema_version": {"type": "string"},
        "scans": {"type": "array"},
    },
    "required": ["generated_at", "tool", "scans"],
}


def format_event_message(event: dict) -> str:
    etype = event.get("type")
    target = event.get("target", "?")
    if etype == "new-port":
        p = event["port"]
        return (
            "NmapX Alert\n\n"
            f"Target: {target}\n"
            "Event: New open port\n\n"
            f"+ {p['port']}/{p['proto']} {p['state']} {p['service']} {p.get('version', '')}"
        ).strip()
    if etype == "version-change":
        p = event["port"]
        old = event["old"]
        return (
            "NmapX Alert\n\n"
            f"Target: {target}\n"
            "Event: Service version changed\n\n"
            f"Port: {p['port']}/{p['proto']} {p['service']}\n"
            f"Before: {old.get('version') or '?'}\n"
            f"After: {p.get('version') or '?'}"
        )
    if etype == "critical-cve":
        p = event["port"]
        cve = event["cve"]
        return (
            "NmapX Alert\n\n"
            f"Target: {target}\n"
            "Event: High-risk CVE\n\n"
            f"Port: {p['port']}/{p['proto']} {p['service']}\n"
            f"CVE: {cve.get('id', '?')}\n"
            f"Severity: {cve.get('severity', '?')} | CVSS: {cve.get('score', '?')}\n"
            f"{cve.get('desc', '')}"
        )
    if etype == "scan-failed":
        return (
            "NmapX Alert\n\n"
            f"Target: {target}\n"
            "Event: Scan failed\n"
            f"Exit code: {event.get('exit_code', '?')}"
        )
    return f"NmapX Alert\n\nTarget: {target}\nEvent: {etype}"


async def send_telegram_notifications(events: List[dict], token: str, chat_id: str, allowed_types: Set[str]):
    if not events or not token or not chat_id:
        return
    if not common.HAS_HTTPX:
        print(f"{common.C.YELLOW}[!] Telegram notifications need httpx. Run: pip install httpx{common.C.RESET}")
        return

    selected = [event for event in events if event.get("type") in allowed_types]
    if not selected:
        return

    url = f"https://api.telegram.org/bot{token}/sendMessage"
    async with common.httpx.AsyncClient(timeout=10.0) as client:
        for event in selected:
            try:
                await client.post(url, json={"chat_id": chat_id, "text": format_event_message(event)})
            except Exception as exc:
                print(f"{common.C.YELLOW}[!] Telegram notification failed: {exc}{common.C.RESET}")


def print_report(result: dict):
    if result.get("error"):
        return

    target = result["target"]
    profile = result["profile"]
    sid = result["sid"]
    ports = result["ports"]
    raw = result.get("raw", "")
    duration = result.get("duration", 0.0)
    resumed = result.get("resumed", False)
    cmd = result.get("cmd", [])

    os_info = ""
    latency = ""
    for line in raw.splitlines():
        if not os_info and ("OS details:" in line or "Aggressive OS guesses:" in line):
            os_info = line.strip()
        if not latency and "Host is up" in line:
            m = re.search(r"\((.+?)\s+latency\)", line)
            if m:
                latency = m.group(1)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"\n{common.C.CYAN}{'═'*65}{common.C.RESET}")
    print(f"  {common.C.BOLD}TARGET   {common.C.RESET}: {common.C.WHITE}{target}{common.C.RESET}"
          + (f"  {common.C.YELLOW}[RESUMED]{common.C.RESET}" if resumed else ""))
    print(f"  {common.C.BOLD}SESSION  {common.C.RESET}: {common.C.DIM}{sid}{common.C.RESET}")
    print(f"  {common.C.BOLD}PROFILE  {common.C.RESET}: {common.C.MAGENTA}{profile}{common.C.RESET}  —  {common.SCAN_PROFILES[profile]['desc']}")
    if cmd:
        print(f"  {common.C.BOLD}COMMAND  {common.C.RESET}: {common.C.DIM}{' '.join(cmd)}{common.C.RESET}")
    if latency:
        print(f"  {common.C.BOLD}LATENCY  {common.C.RESET}: {latency}")
    print(f"  {common.C.BOLD}TIME     {common.C.RESET}: {ts}  ({duration:.1f}s)")
    print(f"{common.C.CYAN}{'═'*65}{common.C.RESET}\n")

    if ports:
        open_c = sum(1 for p in ports if "open" in p["state"])
        filtered_c = sum(1 for p in ports if "filtered" in p["state"])

        print(f"  {common.C.BOLD}{'PORT':<10}{'PROTO':<7}{'STATE':<12}{'SERVICE':<14}{'VERSION':<26}CVEs{common.C.RESET}")
        print(f"  {'─'*78}")

        for p in ports:
            color = common.PORT_COLORS.get(p["state"].split("|")[0], common.C.WHITE)
            cve_cnt = len(p.get("cves", []))
            cve_str = ""
            if cve_cnt:
                sc = CVEEnricher.severity_color(p["cves"][0]["severity"])
                cve_str = f"{sc}[{cve_cnt} CVE{'s' if cve_cnt>1 else ''}]{common.C.RESET}"

            print(
                f"  {common.C.YELLOW}{p['port']:<10}{common.C.RESET}"
                f"{common.C.DIM}{p['proto']:<7}{common.C.RESET}"
                f"{color}{p['state']:<12}{common.C.RESET}"
                f"{common.C.CYAN}{p['service']:<14}{common.C.RESET}"
                f"{common.C.DIM}{p['version'][:25]:<26}{common.C.RESET}"
                f"{cve_str}"
            )

        print(f"\n  {common.C.GREEN}[+] Open: {open_c}  {common.C.YELLOW}[~] Filtered: {filtered_c}{common.C.RESET}")

        cve_ports = [p for p in ports if p.get("cves")]
        if cve_ports:
            print(f"\n  {common.C.RED}{common.C.BOLD}⚠  CVE Findings:{common.C.RESET}")
            for p in cve_ports:
                print(f"\n  {common.C.YELLOW}  {p['port']}/{p['service']} — {p['version']}{common.C.RESET}")
                for cve in p["cves"]:
                    sc = CVEEnricher.severity_color(cve["severity"])
                    print(f"    {sc}{cve['id']}{common.C.RESET}  CVSS:{common.C.BOLD}{cve['score']}{common.C.RESET}"
                          f"  [{sc}{cve['severity']}{common.C.RESET}]  {common.C.DIM}{cve['desc']}{common.C.RESET}")
    else:
        print(f"  {common.C.DIM}No port data (host down or ping-only){common.C.RESET}")

    os_info = result.get('os', '')
    if os_info:
        print(f"\n  {common.C.BOLD}OS:{common.C.RESET} {common.C.MAGENTA}{os_info}{common.C.RESET}")

    print(f"\n{common.C.CYAN}{'═'*65}{common.C.RESET}\n")


def export_json(results: List[dict], path: str):
    data = {
        "$schema": JSON_SCHEMA["$schema"],
        "$id": JSON_SCHEMA["$id"],
        "generated_at": datetime.now().isoformat(),
        "tool": "NmapX v3.0",
        "schema_version": "1.0",
        "scans": results,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    schema_path = os.path.splitext(path)[0] + ".schema.json"
    with open(schema_path, "w", encoding="utf-8") as f:
        json.dump(JSON_SCHEMA, f, indent=2)

    print(f"  {common.C.GREEN}[✓] JSON → {path}{common.C.RESET}")
    print(f"  {common.C.GREEN}[✓] JSON Schema → {schema_path}{common.C.RESET}")


def export_csv(results: List[dict], path: str):
    fieldnames = [
        "target", "profile", "session_id", "os", "port", "proto", "state",
        "service", "version", "cpe", "cve_count", "cve_ids",
        "cve_queries", "cve_confidence", "event_types",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for res in results:
            event_types = ";".join(sorted({e.get("type", "") for e in res.get("events", []) if e.get("type")}))
            for p in res.get("ports", []):
                cves = p.get("cves", []) or []
                writer.writerow({
                    "target": res.get("target", ""),
                    "profile": res.get("profile", ""),
                    "session_id": res.get("sid", ""),
                    "os": res.get("os", ""),
                    "port": p.get("port", ""),
                    "proto": p.get("proto", ""),
                    "state": p.get("state", ""),
                    "service": p.get("service", ""),
                    "version": p.get("version", ""),
                    "cpe": p.get("cpe", ""),
                    "cve_count": len(cves),
                    "cve_ids": ";".join(c.get("id", "") for c in cves),
                    "cve_queries": ";".join(sorted({str(c.get("query", "")) for c in cves if c.get("query")})),
                    "cve_confidence": ";".join(str(c.get("confidence", "")) for c in cves),
                    "event_types": event_types,
                })
    print(f"  {common.C.GREEN}[✓] CSV → {path}{common.C.RESET}")


def export_excel(results: List[dict], path: str):
    if Workbook is None:
        print(f"  {common.C.RED}[!] Excel export requires openpyxl. Run: pip install openpyxl{common.C.RESET}")
        return

    if not path.lower().endswith(".xlsx"):
        path = f"{path}.xlsx"

    wb = Workbook()
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_header = ["Target", "Profile", "Session", "OS", "Open Ports", "CVE Count", "Events"]
    summary_sheet.append(summary_header)
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = yellow_fill

    for res in results:
        event_types = ";".join(sorted({e.get("type", "") for e in res.get("events", []) if e.get("type")}))
        summary_sheet.append([
            res.get("target", ""), res.get("profile", ""), res.get("sid", ""),
            res.get("os", ""), sum(1 for p in res.get("ports", []) if "open" in p.get("state", "")),
            sum(len(p.get("cves", [])) for p in res.get("ports", [])), event_types,
        ])

    ports_sheet = wb.create_sheet("Ports")
    ports_header = ["Target", "Port", "Proto", "State", "Service", "Version", "CPE", "CVE Count"]
    ports_sheet.append(ports_header)
    for cell in ports_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = yellow_fill

    for res in results:
        for p in res.get("ports", []):
            ports_sheet.append([
                res.get("target", ""), p.get("port", ""), p.get("proto", ""),
                p.get("state", ""), p.get("service", ""), p.get("version", ""),
                p.get("cpe", ""), len(p.get("cves", [])),
            ])

    cves_sheet = wb.create_sheet("CVEs")
    cves_header = ["Target", "Port", "CVE", "Severity", "Score", "Query", "Confidence", "Description"]
    cves_sheet.append(cves_header)
    for cell in cves_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = yellow_fill

    for res in results:
        for p in res.get("ports", []):
            for cve in p.get("cves", []):
                cves_sheet.append([
                    res.get("target", ""), p.get("port", ""), cve.get("id", ""),
                    cve.get("severity", ""), cve.get("score", ""), cve.get("query", ""),
                    cve.get("confidence", ""), cve.get("desc", ""),
                ])

    events_sheet = wb.create_sheet("Events")
    events_header = ["Target", "Type", "Port", "Service", "Old Version", "New Version", "CVE ID"]
    events_sheet.append(events_header)
    for cell in events_sheet[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = yellow_fill

    for res in results:
        for event in res.get("events", []):
            port = event.get("port", {}) if isinstance(event.get("port"), dict) else {}
            old = event.get("old", {}) if isinstance(event.get("old"), dict) else {}
            events_sheet.append([
                res.get("target", ""), event.get("type", ""), port.get("port", ""),
                port.get("service", ""), old.get("version", ""), port.get("version", ""),
                event.get("cve", {}).get("id", "") if isinstance(event.get("cve"), dict) else "",
            ])

    wb.save(path)
    print(f"  {common.C.GREEN}[OK] Excel -> {path}{common.C.RESET}")


def print_sessions(db):
    sessions = db.list_sessions()
    if not sessions:
        print(f"\n{common.C.DIM}  No sessions.{common.C.RESET}\n")
        return
    print(f"\n{common.C.BOLD}{common.C.CYAN}  Recent Sessions:{common.C.RESET}")
    print(f"  {'─'*70}")
    print(f"  {common.C.BOLD}{'ID':<14}{'TARGET':<22}{'PROFILE':<12}{'STATUS':<10}STARTED{common.C.RESET}")
    print(f"  {'─'*70}")
    for s in sessions:
        sc = common.C.GREEN if s["status"] == "done" else common.C.YELLOW
        print(f"  {common.C.DIM}{s['id']:<14}{common.C.RESET}{common.C.WHITE}{s['target']:<22}{common.C.RESET}"
              f"{common.C.MAGENTA}{s['profile']:<12}{common.C.RESET}{sc}{s['status']:<10}{common.C.RESET}"
              f"{common.C.DIM}{s['started'][:19]}{common.C.RESET}")
    print()


def print_profiles():
    print(f"\n{common.C.BOLD}{common.C.CYAN}  Scan Profiles:{common.C.RESET}")
    print(f"  {'─'*55}")
    for name, info in common.SCAN_PROFILES.items():
        flags_str = " ".join(info["flags"]) or "(user-defined)"
        print(f"  {common.C.YELLOW}{name:<12}{common.C.RESET} {info['desc']}")
        print(f"  {common.C.DIM}{'':12} {flags_str}{common.C.RESET}")
    print()
